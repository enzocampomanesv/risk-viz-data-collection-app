"""One-off migration: rewrite study_content.xlsx into the new schema.

The overhaul removed the pairwise-comparison and choice-question activities and
merged word association + choice into a single `questions` tab (see
make_template.py / build_content.py). This script carries the existing content
forward:

  - word_prompts        -> questions rows, type=word_prompt (content unchanged)
  - registration        -> unchanged
  - discussion          -> unchanged
  - likert_stimuli      -> unchanged
  - settings            -> only non-pairwise keys kept (currently likert_points)
  - comparison_items    -> dropped (activity retired)
  - choice_questions    -> dropped (activity retired)

It reuses build_workbook() so the output matches a fresh template exactly.
Run once:  python tools/migrate_content.py   (overwrites ../study_content.xlsx)
A timestamped backup of the original is written alongside it first.
"""
import os
import shutil
import time

from openpyxl import load_workbook

from make_template import build_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
XLSX = os.path.join(ROOT, "study_content.xlsx")

# Settings keys that survive the overhaul. Everything else (pairwise_*,
# comparisons_per_group, prep_comparisons) is dropped.
KEEP_SETTINGS = {"likert_points"}


def read_tab(wb, name):
    """Return (list-of-dicts keyed by header, headers) for an existing tab, or
    ([], []) if the tab is absent. Blank rows are skipped."""
    if name not in wb.sheetnames:
        return [], []
    rows = list(wb[name].iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        out.append({headers[j]: (raw[j] if j < len(raw) else None) for j in range(len(headers))})
    return out, headers


def main():
    if not os.path.exists(XLSX):
        raise SystemExit(f"ERROR: {XLSX} not found — nothing to migrate.")

    backup = XLSX.replace(".xlsx", f".pre-migration-{time.strftime('%Y%m%d-%H%M%S')}.xlsx")
    shutil.copy2(XLSX, backup)
    print(f"Backed up original -> {os.path.relpath(backup, ROOT)}")

    wb = load_workbook(XLSX, data_only=True)
    rows_by_tab = {}

    # settings: keep only surviving keys, in their original order.
    settings, _ = read_tab(wb, "settings")
    rows_by_tab["settings"] = [
        {"key": r.get("key"), "value": r.get("value"), "notes": r.get("notes")}
        for r in settings if str(r.get("key")).strip() in KEEP_SETTINGS
    ]

    # registration: straight carry-forward.
    reg, _ = read_tab(wb, "registration")
    rows_by_tab["registration"] = [
        {c: r.get(c) for c in ["field_id", "label", "type", "required", "option_value", "option_text"]}
        for r in reg
    ]

    # word_prompts -> questions (type=word_prompt). One row each, no choices.
    wp, _ = read_tab(wb, "word_prompts")
    rows_by_tab["questions"] = [
        {"id": r.get("id"), "section": r.get("section"), "type": "word_prompt",
         "prompt": r.get("text"), "image": r.get("image"),
         "max_words": r.get("max_words"), "min_words": r.get("min_words")}
        for r in wp
    ]

    # discussion + likert_stimuli: straight carry-forward.
    disc, _ = read_tab(wb, "discussion")
    rows_by_tab["discussion"] = [
        {c: r.get(c) for c in ["id", "text", "image"]} for r in disc
    ]
    lk, _ = read_tab(wb, "likert_stimuli")
    rows_by_tab["likert_stimuli"] = [
        {c: r.get(c) for c in ["id", "title", "body", "image"]} for r in lk
    ]

    build_workbook(rows_by_tab, XLSX)
    print(f"Migrated {os.path.relpath(XLSX, ROOT)}:")
    print(f"  settings   : {len(rows_by_tab['settings'])} row(s) kept")
    print(f"  registration: {len(rows_by_tab['registration'])} row(s)")
    print(f"  questions  : {len(rows_by_tab['questions'])} word_prompt row(s) migrated")
    print(f"  discussion : {len(rows_by_tab['discussion'])} row(s)")
    print(f"  likert_stimuli: {len(rows_by_tab['likert_stimuli'])} row(s)")
    print("Dropped tabs: comparison_items, choice_questions.")
    print("Next: python tools/build_content.py")


if __name__ == "__main__":
    main()
