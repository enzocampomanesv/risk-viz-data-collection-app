"""Generate study_content.xlsx (the team's single editing surface).

Exposes build_workbook(rows_by_tab, path) so the same styled schema is used
both for a blank/example template (this script's main()) and for one-off
content migrations. Run: python tools/make_template.py  (writes ../study_content.xlsx)
"""
import os
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="0E7C86")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Arial", size=11)
WRAP        = Alignment(vertical="top", wrap_text=True)
THIN        = Side(style="thin", color="D9DDE2")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Ordered schema: tab -> columns + display widths. This order is the tab order.
TABS = OrderedDict([
    ("settings",         {"cols": ["key", "value", "notes"],
                          "widths": [26, 18, 56]}),
    ("questions",        {"cols": ["id", "section", "type", "prompt", "image", "profile", "review",
                                   "has_other", "choice_text", "max_words", "min_words", "max_chars"],
                          "widths": [10, 12, 15, 44, 22, 9, 8, 10, 34, 11, 11, 10]}),
    ("stimuli",          {"cols": ["id", "title", "image", "caption"],
                          "widths": [10, 26, 28, 52]}),
    ("wordcloud",        {"cols": ["id", "section", "prompt", "image", "max_words", "min_words", "max_chars"],
                          "widths": [10, 12, 48, 24, 11, 11, 10]}),
])

NOTES = [
    ["How to edit this file"],
    [""],
    ["Each tab is one kind of study content. Edit the rows, save, then run:"],
    ["   python tools/build_content.py     (publishes your changes to the app)"],
    [""],
    ["Rules that matter:"],
    ["- 'id' / question 'id': short and unique WITHIN the tab (q1, st1, d1). Don't reuse."],
    ["- Leave optional cells blank (e.g. 'image', or 'max_words' on a choice question)."],
    ["- Do not rename tabs or the header row (row 1)."],
    [""],
    ["settings:        key/value study-wide options. See the 'notes' column on each row."],
    ["questions:       the whole participant flow, in order. One row PER CHOICE."],
    ["                 Repeat the same 'id' down the rows; the FIRST row carries type/prompt/"],
    ["                 image/profile/review/has_other/word limits, later rows fill 'choice_text'."],
    ["                 type = single_choice | multiple_choice | word_prompt."],
    ["                 - single_choice / multiple_choice: give 2+ 'choice_text' rows. has_other=yes"],
    ["                   appends an 'Other' option with a free-text box (both types)."],
    ["                 - word_prompt: one row, no choices. max_words / min_words bound the number of"],
    ["                   entries; max_chars caps each entry's length (default 30)."],
    ["                 profile=yes (single/multiple choice only): the answer is stored as a"],
    ["                   participant attribute you can group/filter results by (age, gender, role...)."],
    ["                   Single-choice profile = one value; multiple-choice profile = a set (each"],
    ["                   selected value is an overlapping group). Not allowed on word_prompt."],
    ["                 review=yes: after this question, show a 'check your answers' summary of the"],
    ["                   questions so far IN THIS SECTION (a Back to fix, a Confirm to continue)."],
    ["                 'section' = which questionnaire section this question belongs to (a section id"],
    ["                 from config.json, e.g. profile / questions). Blank = the first questionnaire"],
    ["                 section. Questions appear in row order. Rich text in prompt/choices: **bold**."],
    ["discussion + Likert are merged: the host shows a stimulus (figure/gif), the room discusses,"],
    ["                 then the host reveals the scale. Participants score the same three questions"],
    ["                 (set in config.json) on a points-scale (settings: likert_points, odd: 3/5/7)"],
    ["                 labelled only at the ends + middle (settings: anchor_low / anchor_mid / anchor_high)."],
    ["stimuli:         figures shown in the assessment section. One row PER SLIDE; repeat the id"],
    ["                 down the rows to give a stimulus several slides (a carousel). 'title' comes"],
    ["                 from the first row (the Likert score attaches to the id, not the slide)."],
    ["                 Each row is one slide = its image and/or caption. One row = a single image."],
    ["wordcloud:       host-paced word prompts (its own section). One row per prompt: id, section,"],
    ["                 prompt, image (optional figure), max_words / min_words / max_chars. The host"],
    ["                 shows one prompt at a time and clicks Next; participants type words on each."],
    ["settings keys:   likert_points (3/5/7) · anchor_low / anchor_mid / anchor_high (scale labels) ·"],
    ["                 notice copy: notice_*_title / notice_*_body for the between-activity screens"],
    ["                 (see config.json for which section uses which key)."],
]


def build_workbook(rows_by_tab, path):
    wb = Workbook()
    # READ ME FIRST
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.column_dimensions["A"].width = 100
    for r, line in enumerate(NOTES, start=1):
        c = ws.cell(row=r, column=1, value=line[0])
        c.font = Font(name="Arial", bold=(r == 1), size=12 if r == 1 else 11)

    for tab, spec in TABS.items():
        ws = wb.create_sheet(tab)
        cols = spec["cols"]
        for j, (col, w) in enumerate(zip(cols, spec["widths"]), start=1):
            cell = ws.cell(row=1, column=j, value=col)
            cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = BORDER
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = w
        ws.freeze_panes = "A2"
        for i, row in enumerate(rows_by_tab.get(tab, []), start=2):
            for j, col in enumerate(cols, start=1):
                cell = ws.cell(row=i, column=j, value=row.get(col))
                cell.font = BODY_FONT; cell.border = BORDER; cell.alignment = WRAP
    wb.save(path)


# ---- example rows (used for a blank template; migration supplies real rows) ----
EXAMPLE_ROWS = {
    "settings": [
        {"key": "likert_points", "value": 5, "notes": "Points on the scale. Must be odd (3, 5, or 7)."},
        {"key": "anchor_low", "value": "Not at all", "notes": "Label at the low end of the scale."},
        {"key": "anchor_mid", "value": "Partially", "notes": "Label at the midpoint of the scale."},
        {"key": "anchor_high", "value": "Very much", "notes": "Label at the high end of the scale."},
        {"key": "notice_completion1_title", "value": "Great work \u2013 session one is done!",
         "notes": "Shown after the questionnaire; the host moves everyone on."},
        {"key": "notice_completion1_body",
         "value": "You can leave the app and move on to the group activity. We will reconnect back in the app later.",
         "notes": "Body text for the completion-1 screen. Plain text, **bold** allowed."},
        {"key": "notice_welcome_title", "value": "Welcome back!", "notes": "Shown when the host brings the room back."},
        {"key": "notice_welcome_body",
         "value": "Up next are examples of flood communication formats and channels. Review and score each one.",
         "notes": "Body text for the welcome-back screen."},
        {"key": "notice_completion2_title", "value": "Well done!", "notes": "Shown after the assessment section."},
        {"key": "notice_completion2_body",
         "value": "We will now move into our last activity of the day.",
         "notes": "Body text for the completion-2 screen."},
    ],
    "questions": [
        # profile single-choice: stored as a participant attribute for grouping.
        {"id": "age", "section": "profile", "type": "single_choice", "profile": "yes",
         "prompt": "What is your age group?", "choice_text": "18-24"},
        {"id": "age", "choice_text": "25-34"},
        {"id": "age", "choice_text": "35-44"},
        # profile multiple-choice with a review summary right after it.
        {"id": "role", "section": "profile", "type": "multiple_choice", "profile": "yes",
         "review": "yes", "has_other": "yes",
         "prompt": "Do you have any special role in your community?", "choice_text": "Community leader"},
        {"id": "role", "choice_text": "Religious leader"},
        {"id": "role", "choice_text": "Health worker"},
        # regular multiple-choice (a normal response, not a profile attribute).
        {"id": "impacts", "section": "questions", "type": "multiple_choice", "has_other": "yes",
         "prompt": "Which of the following impacts have you experienced?", "choice_text": "Property damage"},
        {"id": "impacts", "choice_text": "Displacement"},
        {"id": "impacts", "choice_text": "Stress"},
        # word_prompt: up to 5 short entries, each capped at 30 characters.
        {"id": "info_recv", "section": "questions", "type": "word_prompt",
         "prompt": "Which flood-related information do you currently receive?",
         "max_words": 5, "min_words": 1, "max_chars": 30},
    ],
    "stimuli": [
        {"id": "st1", "title": "SMS text alert", "image": "images/prompts/sms.png",
         "caption": "A phone SMS: 'FLOOD ALERT: Heavy rain expected tonight. Move to higher ground.'"},
        # A two-slide stimulus: repeat the id, title only on the first row.
        {"id": "st2", "title": "Flood map", "image": "images/prompts/depth_map.jpg",
         "caption": "Flood depth map: colour-coded water depth by area."},
        {"id": "st2", "image": "images/prompts/extent_map.jpg",
         "caption": "Flood extent map: which areas are expected to flood."},
    ],
    "wordcloud": [
        {"id": "wc1", "section": "wordcloud", "prompt": "Through which channels would you prefer to receive flood information?",
         "max_words": 5, "min_words": 1, "max_chars": 30},
        {"id": "wc2", "section": "wordcloud", "prompt": "Which format would be most useful?",
         "max_words": 5, "min_words": 1, "max_chars": 30},
    ],
}

if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "..", "study_content.xlsx")
    build_workbook(EXAMPLE_ROWS, out)
    print("Wrote", os.path.abspath(out))
