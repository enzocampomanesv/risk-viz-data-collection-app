"""Build-time content converter: study_content.xlsx -> config/content.json.

The participant app, control room, and dashboard read this JSON (never Excel),
so the team's single editing surface is the spreadsheet. Run after editing:

    python tools/build_content.py

Validates ids, required columns, settings, groups, and warns on missing images.
Exits non-zero (without writing) on blocking errors.
"""
import json, os, sys
from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
XLSX = os.path.join(ROOT, "study_content.xlsx")
OUT  = os.path.join(ROOT, "config", "content.json")

errors, warnings = [], []

def s(v):
    return "" if v is None else str(v).strip()

def as_bool(v, default=False):
    t = s(v).lower()
    if t in ("yes", "true", "1", "on"): return True
    if t in ("no", "false", "0", "off", ""): return default if t == "" else False
    return default

def as_int(v, default, where):
    try:
        return int(float(s(v)))
    except Exception:
        if s(v) != "":
            errors.append(f"{where}: '{s(v)}' is not a whole number.")
        return default

def read_sheet(wb, name):
    if name not in wb.sheetnames:
        errors.append(f"[{name}] tab is missing from the workbook.")
        return [], []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [s(h) for h in rows[0]]
    out = []
    for i, raw in enumerate(rows[1:], start=2):
        if all(c is None or s(c) == "" for c in raw):
            continue
        rec = {headers[j]: (raw[j] if j < len(raw) else None) for j in range(len(headers))}
        rec["__row"] = i
        out.append(rec)
    return out, headers

def questionnaire_section_ids():
    """Ordered ids of questionnaire sections in config.json (for scoping
    questions). First entry is the default home for un-tagged questions."""
    try:
        cfg = json.load(open(os.path.join(ROOT, "config", "config.json"), encoding="utf-8"))
        return [s["id"] for s in cfg.get("sections", []) if s.get("type") == "questionnaire"]
    except Exception:
        return []

def wordcloud_section_ids():
    """Ordered ids of wordcloud sections in config.json (for scoping wordcloud
    prompts). First entry is the default home for un-tagged prompts."""
    try:
        cfg = json.load(open(os.path.join(ROOT, "config", "config.json"), encoding="utf-8"))
        return [s["id"] for s in cfg.get("sections", []) if s.get("type") == "wordcloud"]
    except Exception:
        return []

PROMPT_FOLDER = "images/prompts"

def resolve_image(raw):
    """'image' cells (questions / discussion / likert_stimuli) accept either a
    bare filename (resolved against PROMPT_FOLDER) or a full path (containing
    '/'), kept as-is for backward compatibility with rows already written as
    images/prompts/xx.jpg."""
    v = s(raw)
    if not v:
        return None
    return v if "/" in v else f"{PROMPT_FOLDER}/{v}"

def check_image(path, where):
    if path and not os.path.exists(os.path.join(ROOT, path)):
        warnings.append(f"{where}: image not found yet: {path}")

def dup_check(seen, rid, where):
    if rid in seen:
        errors.append(f"{where}: duplicate id '{rid}'.")
    seen.add(rid)

def main():
    if not os.path.exists(XLSX):
        sys.exit(f"ERROR: {XLSX} not found. Run tools/make_template.py first.")
    wb = load_workbook(XLSX, data_only=True)
    content = {}

    # ---------- settings ----------
    recs, _ = read_sheet(wb, "settings")
    kv = {s(r.get("key")): r.get("value") for r in recs}
    settings = {
        "likert_points": as_int(kv.get("likert_points"), 5, "[settings] likert_points"),
        "anchor_low":  s(kv.get("anchor_low"))  or "Not at all",
        "anchor_mid":  s(kv.get("anchor_mid"))  or "Partially",
        "anchor_high": s(kv.get("anchor_high")) or "Very much",
    }
    if settings["likert_points"] < 3 or settings["likert_points"] % 2 == 0:
        errors.append("[settings] likert_points must be an odd number >= 3 (e.g. 3, 5, or 7) "
                      "so the scale has a true midpoint.")
    # Free-text notice copy: any key starting with notice_ is carried through as-is.
    for k, v in kv.items():
        if k.startswith("notice_"):
            settings[k] = s(v)
    content["settings"] = settings

    # ---------- questions (questionnaire content) ----------
    # Long format: one row per choice, the question id repeated down the rows.
    # The FIRST row for an id is its header (type/prompt/image/profile/review/
    # has_other/word limits); every row (header included) may carry one
    # choice_text. A repeated id means "same question", not a duplicate.
    #
    # profile=yes marks a single/multiple-choice question whose answer is stored
    # as a participant attribute (for grouping/filtering results); participant_
    # fields are derived from these in config-loader.js, so there's no separate
    # registration tab. review=yes marks a "check your answers" summary point.
    recs, _ = read_sheet(wb, "questions")
    q_secs = questionnaire_section_ids()
    default_q_sec = q_secs[0] if q_secs else None
    TYPE_ALIASES = {
        "single_choice": "single_choice", "single-choice": "single_choice", "single": "single_choice",
        "multiple_choice": "multiple_choice", "multiple-choice": "multiple_choice",
        "multi_choice": "multiple_choice", "multi-choice": "multiple_choice", "multi": "multiple_choice",
        "word_prompt": "word_prompt", "word-prompt": "word_prompt", "word": "word_prompt",
    }
    CHOICE_TYPES = ("single_choice", "multiple_choice")
    order, byid = [], {}
    for r in recs:
        rid, row = s(r.get("id")), r["__row"]
        if not rid:
            errors.append(f"[questions row {row}] missing id "
                          f"(every row, including extra choice rows, needs the question id).")
            continue
        if rid not in byid:
            raw_type = s(r.get("type")).lower()
            qtype = TYPE_ALIASES.get(raw_type)
            if qtype is None:
                errors.append(f"[questions row {row}] question '{rid}' has invalid or missing type "
                              f"'{s(r.get('type'))}' (use single_choice | multiple_choice | word_prompt).")
            byid[rid] = {"id": rid, "row": row, "type": qtype,
                         "section": s(r.get("section")) or default_q_sec,
                         "prompt": s(r.get("prompt")), "image": resolve_image(r.get("image")),
                         "profile": as_bool(r.get("profile"), False),
                         "review": as_bool(r.get("review"), False),
                         "has_other": as_bool(r.get("has_other"), False),
                         "max_words_raw": r.get("max_words"), "min_words_raw": r.get("min_words"),
                         "max_chars_raw": r.get("max_chars"),
                         "choices": []}
            order.append(rid)
        ct = s(r.get("choice_text"))
        if ct:
            byid[rid]["choices"].append(ct)

    clean = []
    for rid in order:
        q = byid[rid]
        where = f"[questions '{rid}']"
        if q["type"] is None:
            continue  # invalid type already reported
        if q_secs and q["section"] not in q_secs:
            errors.append(f"{where} section '{q['section']}' is not a questionnaire section "
                          f"(valid: {', '.join(q_secs)}).")
        if not q["prompt"]:
            errors.append(f"{where} missing prompt.")
        check_image(q["image"], where)
        rec = {"id": rid, "section": q["section"], "type": q["type"],
               "prompt": q["prompt"], "image": q["image"]}
        if q["review"]:
            rec["review"] = True
        if q["type"] in CHOICE_TYPES:
            if len(q["choices"]) < 2:
                errors.append(f"{where} needs at least 2 choices (has {len(q['choices'])}). "
                              f"Add a row per choice with the same id.")
            rec["choices"] = q["choices"]
            rec["has_other"] = q["has_other"]
            if q["profile"]:
                rec["profile"] = True
        else:  # word_prompt
            if q["profile"]:
                errors.append(f"{where} profile=yes is only allowed on single_choice / "
                              f"multiple_choice questions (a word_prompt can't be a group key).")
            maxw = as_int(q["max_words_raw"], 5, f"{where} max_words")
            minw = as_int(q["min_words_raw"], 1, f"{where} min_words")
            maxc = as_int(q["max_chars_raw"], 30, f"{where} max_chars")
            if minw < 1: errors.append(f"{where} min_words must be >= 1.")
            if maxw < minw: errors.append(f"{where} max_words < min_words.")
            if maxc < 1: errors.append(f"{where} max_chars must be >= 1.")
            if q["choices"]:
                warnings.append(f"{where} is a word_prompt but has choice_text rows — they are ignored.")
            if q["has_other"]:
                warnings.append(f"{where} has_other is ignored on a word_prompt.")
            rec["max_words"] = maxw
            rec["min_words"] = minw
            rec["max_chars"] = maxc
        clean.append(rec)
    content["questions"] = clean

    # ---------- stimuli (assessment figures) ----------
    # Long format: one row per SLIDE, the stimulus id repeated down the rows.
    # A stimulus is one id + one title (the Likert attaches to the id) holding a
    # carousel of slides; each row contributes a slide = its image and/or caption.
    # One row for an id = a single slide (behaves like before).
    recs, _ = read_sheet(wb, "stimuli")
    order, byid = [], {}
    for r in recs:
        rid, row = s(r.get("id")), r["__row"]
        if not rid:
            errors.append(f"[stimuli row {row}] missing id."); continue
        if rid not in byid:
            byid[rid] = {"id": rid, "title": None, "slides": []}
            order.append(rid)
        if s(r.get("title")) and not byid[rid]["title"]:
            byid[rid]["title"] = s(r.get("title"))
        img = resolve_image(r.get("image"))
        cap = s(r.get("caption")) or None
        check_image(img, f"[stimuli row {row}]")
        if img or cap:
            byid[rid]["slides"].append({"image": img, "caption": cap})
    clean = []
    for rid in order:
        stim = byid[rid]
        if not stim["slides"]:
            errors.append(f"[stimuli '{rid}'] needs at least one row with an image or caption.")
            continue
        # Top-level image/caption mirror the first slide for any single-slide
        # consumer (control-room preview, dashboard labels).
        stim["image"] = stim["slides"][0]["image"]
        stim["caption"] = stim["slides"][0]["caption"]
        clean.append(stim)
    content["stimuli"] = clean

    # ---------- wordcloud (host-paced word prompts) ----------
    recs, _ = read_sheet(wb, "wordcloud")
    wc_secs = wordcloud_section_ids()
    default_wc_sec = wc_secs[0] if wc_secs else None
    seen, clean = set(), []
    for r in recs:
        rid, row = s(r.get("id")), r["__row"]
        if not rid: errors.append(f"[wordcloud row {row}] missing id."); continue
        dup_check(seen, rid, f"[wordcloud row {row}]")
        prompt = s(r.get("prompt"))
        if not prompt: errors.append(f"[wordcloud row {row}] missing prompt.")
        img = resolve_image(r.get("image")); check_image(img, f"[wordcloud row {row}]")
        maxw = as_int(r.get("max_words"), 5, f"[wordcloud row {row}] max_words")
        minw = as_int(r.get("min_words"), 1, f"[wordcloud row {row}] min_words")
        maxc = as_int(r.get("max_chars"), 30, f"[wordcloud row {row}] max_chars")
        if minw < 1: errors.append(f"[wordcloud row {row}] min_words must be >= 1.")
        if maxw < minw: errors.append(f"[wordcloud row {row}] max_words < min_words.")
        if maxc < 1: errors.append(f"[wordcloud row {row}] max_chars must be >= 1.")
        sec = s(r.get("section")) or default_wc_sec
        if wc_secs and sec not in wc_secs:
            errors.append(f"[wordcloud row {row}] section '{sec}' is not a wordcloud section "
                          f"(valid: {', '.join(wc_secs)}).")
        clean.append({"id": rid, "section": sec, "prompt": prompt, "image": img,
                      "max_words": maxw, "min_words": minw, "max_chars": maxc})
    content["wordcloud"] = clean

    if errors:
        print("BUILD FAILED — fix these and re-run:\n  - " + "\n  - ".join(errors))
        if warnings: print("\nWarnings:\n  - " + "\n  - ".join(warnings))
        sys.exit(1)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    json.dump(content, open(OUT, "w", encoding="utf-8"), indent=2, ensure_ascii=False)
    counts = ", ".join(f"{k}={len(v)}" for k, v in content.items() if isinstance(v, list))
    print(f"OK — wrote {os.path.relpath(OUT, ROOT)}  ({counts})")
    qtypes = {}
    for q in content["questions"]:
        qtypes[q["type"]] = qtypes.get(q["type"], 0) + 1
    qsummary = ", ".join(f"{t}={n}" for t, n in qtypes.items()) or "none"
    nprofile = sum(1 for q in content["questions"] if q.get("profile"))
    print(f"   questions: {qsummary} | profile={nprofile} | likert={settings['likert_points']}pt")
    if warnings:
        print("Warnings (non-blocking):\n  - " + "\n  - ".join(warnings))

if __name__ == "__main__":
    main()
